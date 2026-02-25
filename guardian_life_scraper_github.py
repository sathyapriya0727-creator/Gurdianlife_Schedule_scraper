# -*- coding: utf-8 -*-
"""
Guardian Life Career Scraper - GitHub Actions Version
Runs automatically on scheduled days via GitHub Actions
Schedule: Monday, Wednesday, Friday at 8:00 AM IST (2:30 AM UTC)
"""

from bs4 import BeautifulSoup
import requests
from tqdm import tqdm
import pandas as pd
import os
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import json
import logging
import time
import sys

# ============================================================================
# CONFIGURATION
# ============================================================================

OUTPUT_FOLDER = 'output'
LOG_FOLDER = 'logs'

EXPORT_CONFIG = {
    'save_excel': True,
    'save_csv': True,
    'save_json': True,
}

MAX_JOBS = 1000        # Safety limit - max jobs to collect
REQUEST_DELAY = 1.0    # Seconds between requests (be polite)
MAX_RETRIES = 3        # Retry failed requests this many times

# ============================================================================
# SETUP AND UTILITY FUNCTIONS
# ============================================================================

def setup_folders():
    """Create necessary folders if they don't exist"""
    for folder in [OUTPUT_FOLDER, LOG_FOLDER]:
        os.makedirs(folder, exist_ok=True)

def setup_logging():
    """Setup logging to both file and console"""
    log_path = os.path.join(LOG_FOLDER, f'scraper_{get_timestamp()}.log')
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_path, encoding='utf-8'),
            logging.StreamHandler(sys.stdout)
        ]
    )
    return logging.getLogger(__name__)

def get_timestamp():
    """Current timestamp in IST"""
    return (datetime.utcnow() + timedelta(hours=5, minutes=30)).strftime('%Y-%m-%d_%H-%M-%S')

def get_date_only():
    """Current date in IST"""
    return (datetime.utcnow() + timedelta(hours=5, minutes=30)).strftime('%Y-%m-%d')

def save_run_history(status, records_count=0, error=None):
    """Append run result to history JSON"""
    history_file = os.path.join(LOG_FOLDER, 'run_history.json')
    entry = {
        'timestamp': get_timestamp(),
        'date': get_date_only(),
        'status': status,
        'records_scraped': records_count,
        'error': str(error) if error else None
    }
    history = []
    if os.path.exists(history_file):
        try:
            with open(history_file, 'r') as f:
                history = json.load(f)
        except Exception:
            history = []
    history.append(entry)
    with open(history_file, 'w') as f:
        json.dump(history, f, indent=2)

# ============================================================================
# HTTP SESSION SETUP
# ============================================================================

def create_session():
    """
    Create a requests Session.

    IMPORTANT — COOKIES EXPIRE PERIODICALLY:
    If the scraper returns 0 jobs or authentication errors, you need to
    refresh the cookies below. Here's how:
      1. Open https://guardianlife.wd5.myworkdayjobs.com in Chrome
      2. Press F12 → Network tab
      3. Click any job or apply a filter
      4. Find a POST request to /jobs
      5. Right-click → Copy → Copy as cURL
      6. Extract the cookie string and update COOKIES below
    """

    COOKIES = {
        'PLAY_SESSION': 'ed1dd99ad8309df39955ca94d4339425751d3c11-guardianlife_pSessionId=iruc1382sugs7n6a53hn84di5b&instance=vps-prod-ie1iirjt.prod-vps.pr503.cust.pdx.wd',
        'wd-browser-id': 'f2fc685c-8c80-4fae-8bea-36a27ee20867',
        'CALYPSO_CSRF_TOKEN': 'd1382620-21d7-4837-8b5b-310c71c84649',
        'wday_vps_cookie': '132946954.53810.0000',
        'timezoneOffset': '-330',
    }

    HEADERS = {
        'accept': 'application/json',
        'accept-language': 'en-US',
        'content-type': 'application/json',
        'user-agent': (
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
            'AppleWebKit/537.36 (KHTML, like Gecko) '
            'Chrome/138.0.0.0 Safari/537.36'
        ),
    }

    session = requests.Session()
    session.cookies.update(COOKIES)
    session.headers.update(HEADERS)
    return session

# ============================================================================
# SCRAPING FUNCTIONS
# ============================================================================

def fetch_with_retry(session, method, url, retries=MAX_RETRIES, **kwargs):
    """Make HTTP request with automatic retry on failure"""
    for attempt in range(1, retries + 1):
        try:
            response = getattr(session, method)(url, timeout=30, **kwargs)
            response.raise_for_status()
            return response
        except requests.exceptions.RequestException as e:
            logger.warning(f"Attempt {attempt}/{retries} failed for {url}: {e}")
            if attempt < retries:
                time.sleep(2 * attempt)  # Backoff: 2s, 4s, 6s
    return None

def fetch_job_list(session, offset):
    """Fetch one page of job listings"""
    payload = {
        'appliedFacets': {},
        'limit': 20,
        'offset': offset,
        'searchText': '',
    }
    url = 'https://guardianlife.wd5.myworkdayjobs.com/wday/cxs/guardianlife/Guardian-Life-Careers/jobs'
    response = fetch_with_retry(session, 'post', url, json=payload)
    if response:
        try:
            return response.json()
        except Exception as e:
            logger.error(f"Failed to parse job list JSON at offset {offset}: {e}")
    return {'jobPostings': []}

def fetch_job_details(session, external_path):
    """Fetch full details for a single job posting"""
    url = f'https://guardianlife.wd5.myworkdayjobs.com/wday/cxs/guardianlife/Guardian-Life-Careers{external_path}'
    response = fetch_with_retry(session, 'get', url)
    if response:
        try:
            return response.json()
        except Exception as e:
            logger.error(f"Failed to parse job details JSON for {external_path}: {e}")
    return {}

def html_to_text(html_string):
    """Strip HTML tags and return clean plain text"""
    if not html_string or (isinstance(html_string, float) and pd.isna(html_string)):
        return ""
    soup = BeautifulSoup(str(html_string), 'html.parser')
    return soup.get_text(separator=' ', strip=True)

def clean_list_field(value):
    """Convert list fields (like additionalLocations) to a readable string"""
    if isinstance(value, list):
        return ', '.join(str(v) for v in value if v)
    return value

def scrape_jobs():
    """
    Main scraping logic:
    1. Collect all job listing summaries (paginated)
    2. Fetch full details for each job
    3. Merge and clean into a DataFrame
    """
    logger.info("Creating HTTP session...")
    session = create_session()

    # --- Phase 1: Collect job listing summaries ---
    logger.info("Phase 1: Collecting job listings...")
    all_postings = []

    for offset in tqdm(range(0, MAX_JOBS, 20), desc='Fetching pages'):
        result = fetch_job_list(session, offset)
        postings = result.get('jobPostings', [])

        if not postings:
            logger.info(f"No more jobs found at offset {offset}. Stopping.")
            break

        all_postings.extend(postings)
        logger.info(f"  Offset {offset}: +{len(postings)} jobs (total: {len(all_postings)})")
        time.sleep(REQUEST_DELAY)

    if not all_postings:
        logger.warning("No job postings collected. Check if cookies are still valid.")
        return None

    logger.info(f"Total listings collected: {len(all_postings)}")

    # Deduplicate by bulletFields
    listings_df = pd.json_normalize(all_postings)
    if 'bulletFields' in listings_df.columns:
        listings_df = listings_df.drop_duplicates(subset=['bulletFields'])
    logger.info(f"After deduplication: {len(listings_df)} unique jobs")

    # --- Phase 2: Fetch full details for each job ---
    logger.info("Phase 2: Fetching job details...")
    all_details = []

    for path in tqdm(listings_df['externalPath'], desc='Fetching details'):
        detail = fetch_job_details(session, path)
        if detail:
            detail['_externalPath'] = path
            all_details.append(detail)
        time.sleep(REQUEST_DELAY)

    if not all_details:
        logger.warning("No job details collected.")
        return None

    details_df = pd.json_normalize(all_details)

    # --- Phase 3: Clean and merge ---
    logger.info("Phase 3: Merging and cleaning data...")

    # Clean HTML from job description
    if 'jobPostingInfo.jobDescription' in details_df.columns:
        details_df['jobPostingInfo.jobDescription'] = (
            details_df['jobPostingInfo.jobDescription'].apply(html_to_text)
        )

    # Merge listings + details
    final_df = pd.merge(
        listings_df,
        details_df,
        left_on='externalPath',
        right_on='_externalPath',
        how='left'
    )

    # Map raw column names to human-readable names
    column_mapping = {
        'jobPostingInfo.title': 'Job Title',
        'jobPostingInfo.jobDescription': 'Job Description',
        'jobPostingInfo.location': 'Location',
        'jobPostingInfo.additionalLocations': 'Additional Locations',
        'jobPostingInfo.startDate': 'Posted Date',
        'jobPostingInfo.jobReqId': 'Job ID',
        'jobPostingInfo.remoteType': 'Remote Type',
        'jobPostingInfo.externalUrl': 'Application URL',
    }

    available = [col for col in column_mapping if col in final_df.columns]
    final_df = final_df[available].rename(columns=column_mapping)

    # Clean list-type fields
    if 'Additional Locations' in final_df.columns:
        final_df['Additional Locations'] = final_df['Additional Locations'].apply(clean_list_field)

    # Add scrape metadata
    final_df.insert(0, 'Scraped Date', get_date_only())

    logger.info(f"Final dataset: {len(final_df)} jobs, {len(final_df.columns)} columns")
    return final_df

# ============================================================================
# EXCEL FORMATTING
# ============================================================================

def format_excel(file_path):
    """Apply professional formatting to the Excel output"""
    wb = load_workbook(file_path)
    ws = wb.active

    # Header style
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Cell style
    cell_font = Font(name='Arial', size=10)
    cell_align = Alignment(vertical='top', wrap_text=True)

    # Alternating row color
    alt_fill = PatternFill(start_color='EBF3FB', end_color='EBF3FB', fill_type='solid')

    thin = Side(style='thin', color='BDD7EE')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Apply header row formatting
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    # Apply data row formatting
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        fill = alt_fill if row_idx % 2 == 0 else None
        for cell in row:
            cell.font = cell_font
            cell.alignment = cell_align
            cell.border = border
            if fill:
                cell.fill = fill

    # Column widths (adjust as needed)
    col_widths = {
        'A': 14,  # Scraped Date
        'B': 38,  # Job Title
        'C': 65,  # Job Description
        'D': 22,  # Location
        'E': 22,  # Additional Locations
        'F': 14,  # Posted Date
        'G': 16,  # Job ID
        'H': 16,  # Remote Type
        'I': 55,  # Application URL
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = 'A2'
    ws.row_dimensions[1].height = 32
    ws.auto_filter.ref = ws.dimensions

    wb.save(file_path)
    logger.info(f"Excel formatting applied: {file_path}")

# ============================================================================
# EXPORT FUNCTIONS
# ============================================================================

def export_data(df):
    """Save DataFrame to configured file formats"""
    date_str = get_date_only()
    exported = []

    if EXPORT_CONFIG['save_excel']:
        path = os.path.join(OUTPUT_FOLDER, f'GuardianLife_Jobs_{date_str}.xlsx')
        df.to_excel(path, index=False, engine='openpyxl')
        format_excel(path)
        exported.append(path)
        logger.info(f"Excel saved: {os.path.basename(path)}")

    if EXPORT_CONFIG['save_csv']:
        path = os.path.join(OUTPUT_FOLDER, f'GuardianLife_Jobs_{date_str}.csv')
        df.to_csv(path, index=False, encoding='utf-8-sig')
        exported.append(path)
        logger.info(f"CSV saved: {os.path.basename(path)}")

    if EXPORT_CONFIG['save_json']:
        path = os.path.join(OUTPUT_FOLDER, f'GuardianLife_Jobs_{date_str}.json')
        df.to_json(path, orient='records', indent=2, force_ascii=False)
        exported.append(path)
        logger.info(f"JSON saved: {os.path.basename(path)}")

    return exported

# ============================================================================
# MAIN
# ============================================================================

def main():
    global logger

    setup_folders()
    logger = setup_logging()

    logger.info("=" * 70)
    logger.info("Guardian Life Career Scraper — GitHub Actions Auto Run")
    logger.info(f"Run time (IST): {get_timestamp().replace('_', ' ')}")
    logger.info("=" * 70)

    try:
        df = scrape_jobs()

        if df is None or df.empty:
            logger.warning("No data scraped. Verify cookies are still valid.")
            save_run_history('no_data')
            print("\n⚠️  No jobs found — cookies may have expired.")
            sys.exit(1)

        exported_files = export_data(df)
        save_run_history('success', len(df))

        logger.info("=" * 70)
        logger.info(f"✅ Scraping completed successfully!")
        logger.info(f"   Total jobs: {len(df)}")
        logger.info(f"   Files saved: {len(exported_files)}")
        for f in exported_files:
            logger.info(f"   → {os.path.basename(f)}")
        logger.info("=" * 70)

        print(f"\n{'=' * 70}")
        print(f"✅  SUCCESS — {len(df)} jobs scraped")
        print(f"{'=' * 70}")
        for f in exported_files:
            print(f"   📄 {os.path.basename(f)}")
        print(f"{'=' * 70}\n")

        return df

    except Exception as e:
        logger.error(f"Fatal error: {e}", exc_info=True)
        save_run_history('error', error=e)
        print(f"\n❌ Error: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
